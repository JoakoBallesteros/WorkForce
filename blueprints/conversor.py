from flask import (
    Blueprint, render_template, request,
    flash, send_file, url_for, current_app, session, redirect
)
import pandas as pd
import os
import random
from datetime import timedelta
from openpyxl.styles import PatternFill, Font

conversor_bp = Blueprint('conversor', __name__, template_folder='templates')

# Lista de servicios: quitamos los que no existen y dejamos "ALL" al principio
SERVICES = [
    'ALL',
    'Sop_Conectividad', 'Sop_Flow', 'Esp_CATV',
    'Esp_Movil', 'Esp_XDSL', 'Digital', 'CBS'
]

DAY_NAMES = {
    0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves',
    4: 'Viernes', 5: 'Sábado', 6: 'Domingo'
}


@conversor_bp.route('/conversor', methods=['GET', 'POST'])
def conversor():
    download_url = None

    if request.method == 'POST':
        servicio  = request.form.get('servicio')
        prog_file = request.files.get('prog_file')

        # Validaciones básicas
        if not servicio or not prog_file:
            flash('Selecciona un servicio y sube un archivo de programación.', 'warning')
            return render_template('conversor.html', title='Conversor', services=SERVICES)

        upload_dir = current_app.config.get('UPLOAD_FOLDER', os.getcwd())
        in_path    = os.path.join(upload_dir, 'input_prog.xlsx')
        prog_file.save(in_path)

        # -----------------------------------------------------------------------------------------
        # 1) Construir el DataFrame “df” destinando la columna SERVICIO según corresponda
        # -----------------------------------------------------------------------------------------
        if servicio == 'ALL':
            # Leemos todas las hojas en un dict
            all_sheets: dict = pd.read_excel(in_path, sheet_name=None)
            partes = []
            for sheet_name, df_sheet in all_sheets.items():
                # Nos quedamos solo con las hojas que contengan "Simulacion" en el nombre
                if 'Simulacion' in sheet_name:
                    # Extraemos el prefijo (por ejemplo "Sop_Conectividad" de
                    # "Sop_Conectividad_Simulacion")
                    prefijo = sheet_name.split('_Simulacion')[0]
                    df_copy = df_sheet.copy()
                    df_copy['SERVICIO'] = prefijo
                    partes.append(df_copy)

            if not partes:
                flash("No se encontró ninguna hoja de 'Simulacion' en el archivo ALL.", 'danger')
                return render_template('conversor.html', title='Conversor', services=SERVICES)

            # Concatenamos todas las porciones y obtenemos un único DataFrame
            df = pd.concat(partes, ignore_index=True)

        else:
            # Caso servicio individual: leemos la hoja por defecto y filtramos por columna 'SERVICIO'
            df = pd.read_excel(in_path)
            if 'SERVICIO' in df.columns:
                key = servicio.split('_')[-1]
                df = df[df['SERVICIO'].str.contains(key, case=False, na=False)]

            # Si no existía la columna 'SERVICIO', la creamos con el valor del servicio seleccionado
            if 'SERVICIO' not in df.columns:
                df['SERVICIO'] = servicio

        # -----------------------------------------------------------------------------------------
        # 2) Expandir “Nombres_Presentes” en filas individuales, creando la columna “Nombre”
        # -----------------------------------------------------------------------------------------
        df = df.dropna(subset=['Nombres_Presentes'])
        records = []
        for _, row in df.iterrows():
            for raw in str(row['Nombres_Presentes']).split(';'):
                name = raw.strip().upper()
                if not name:
                    continue
                # Si viene en formato "APELLIDO, NOMBRE", invertimos a "NOMBRE APELLIDO"
                if ',' in name:
                    ape, nom = [p.strip() for p in name.split(',', 1)]
                    name = f"{nom} {ape}"
                new_row = row.copy()
                new_row['Nombre'] = name
                records.append(new_row)
        df = pd.DataFrame(records)

        # -----------------------------------------------------------------------------------------
        # 3) Cargar la nómina para obtener DNI, SUPERVISOR e INGRESO
        # -----------------------------------------------------------------------------------------
        nomina_path = session.get('nomina_path')
        if not nomina_path or not os.path.exists(nomina_path):
            flash("No se encontró la nómina cargada en sesión.", 'danger')
            return render_template('conversor.html', title='Conversor', services=SERVICES)

        df_nom = pd.read_excel(nomina_path)
        df_nom.columns = df_nom.columns.str.strip()
        required = ('NOMBRE', 'DNI', 'SUPERIOR', 'INGRESO')
        if not all(c in df_nom.columns for c in required):
            flash("La nómina debe tener columnas 'NOMBRE', 'DNI', 'SUPERIOR' e 'INGRESO'.", 'danger')
            return render_template('conversor.html', title='Conversor', services=SERVICES)

        df_nom = df_nom[['NOMBRE', 'DNI', 'SUPERIOR', 'INGRESO']].rename(
            columns={'NOMBRE': 'Nombre', 'DNI': 'DNI', 'INGRESO': 'Ingreso'}
        )
        df_nom['Nombre']  = df_nom['Nombre'].str.upper().str.strip()
        df_nom['DNI']     = df_nom['DNI'].astype(str).str.strip()
        df_nom['Ingreso'] = pd.to_datetime(
            df_nom['Ingreso'].astype(str),
            format='%H:%M:%S',
            errors='coerce'
        ).dt.time

        # -----------------------------------------------------------------------------------------
        # 4) Merge de programación + nómina
        # -----------------------------------------------------------------------------------------
        df = df.merge(df_nom, on='Nombre', how='left')
        # Si falta DNI, SUPERVISOR o Ingreso, avisamos
        if df[['DNI', 'SUPERIOR', 'Ingreso']].isnull().any().any():
            flash("Revisa que los nombres coincidan y que 'DNI','SUPERIOR','INGRESO' estén presentes.", 'danger')
            return render_template('conversor.html', title='Conversor', services=SERVICES)

        # -----------------------------------------------------------------------------------------
        # 5) Procesar fechas e intervalos
        # -----------------------------------------------------------------------------------------
        df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce').dt.date
        df = df.dropna(subset=['Fecha'])
        df['Intervalo_dt'] = pd.to_datetime(
            df['Intervalo'], format='%H:%M', errors='coerce'
        ).dt.time
        df['Intervalo']    = df['Intervalo_dt'].astype(str)
        df['Semana']       = df['Fecha'].apply(lambda d: d - timedelta(days=d.weekday()))
        df = df.sort_values(['Semana', 'Fecha', 'Intervalo_dt'])

        # -----------------------------------------------------------------------------------------
        # 6) Generar Excel de salida: para cada semana agrupamos y pivotamos
        #    Incluimos la columna “SERVICIO” en el resultado final, justo después de “Nombre”.
        #    Agregamos algo de variación al elegir el break (random entre opciones válidas).
        # -----------------------------------------------------------------------------------------
        file_name = f'convertido_tabs_{servicio}.xlsx'
        out_path  = os.path.join(upload_dir, file_name)
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            header_blue   = PatternFill('solid', fgColor='538DD5')
            header_yellow = PatternFill('solid', fgColor='FFC000')
            font_bold     = Font(color='000000', bold=True)

            # Para cada grupo de la misma semana:
            for semana, group in df.groupby('Semana'):
                hoja = semana.strftime('Sem %Y-%m-%d')
                tmp  = group.copy()

                # Creamos columnas auxiliares
                tmp['Dia_Num']  = tmp['Fecha'].apply(lambda d: d.weekday())
                tmp['Presente'] = 1

                # Pivot: filas = Nombre, columnas = día de la semana, valores = presencia
                pivot = tmp.pivot_table(
                    index='Nombre', columns='Dia_Num',
                    values='Presente', aggfunc='sum', fill_value=0
                )
                # Nos aseguramos de que existan las 7 columnas (0 a 6)
                for i in range(7):
                    pivot[i] = pivot.get(i, 0)
                pivot = pivot.reindex(columns=range(7), fill_value=0)
                pivot.rename(columns=DAY_NAMES, inplace=True)
                pivot = pivot.reset_index()

                # Convertimos 0→"Franco", >0→1
                for col in DAY_NAMES.values():
                    pivot[col] = pivot[col].apply(lambda v: 1 if v > 0 else 'Franco')

                # ---------------------------------------------------------------------------------
                # Ahora creamos “map_df” para: 
                #  • extraer SUPERVISOR (primera ocurrencia por Nombre)
                #  • extraer DNI e Ingreso de df_nom
                #  • extraer SERVICIO (primera ocurrencia por Nombre)
                # ---------------------------------------------------------------------------------
                map_df = tmp.groupby('Nombre').agg({
                    'SUPERIOR': 'first',
                    'SERVICIO': 'first'
                }).reset_index()

                map_df = map_df.merge(
                    df_nom[['Nombre', 'DNI', 'Ingreso']],
                    on='Nombre', how='left'
                )
                map_df['Intervalo'] = map_df['Ingreso'].apply(
                    lambda t: t.strftime('%H:%M') if pd.notnull(t) else ''
                )

                pivot = pivot.merge(
                    map_df[['Nombre', 'DNI', 'SUPERIOR', 'SERVICIO', 'Intervalo']],
                    on='Nombre', how='left'
                )

                break_cols = []
                for dnum, dname in DAY_NAMES.items():
                    colb = f'Break_{dname}'
                    break_cols.append(colb)
                    brk = {}
                    day_rows = tmp[tmp['Dia_Num'] == dnum]
                    for nm, grp in day_rows.groupby('Nombre'):
                        horas = sorted(grp['Intervalo_dt'])
                        if not horas:
                            continue
                        if len(horas) < 3:
                            # Pocas franjas: tomamos la mediana (sin variación)
                            med = horas[len(horas)//2]
                            brk[nm] = med.strftime('%H:%M')
                        else:
                            # Calculamos rango válido (entre primera+2h y última−2h)
                            inicio = pd.to_datetime(horas[0].strftime('%H:%M')) + timedelta(hours=2)
                            fin    = pd.to_datetime(horas[-1].strftime('%H:%M')) - timedelta(hours=2)
                            poss = [
                                pd.to_datetime(h.strftime('%H:%M'))
                                for h in horas
                                if inicio <= pd.to_datetime(h.strftime('%H:%M')) <= fin
                            ]
                            if poss:
                                # Elegimos uno al azar entre las opciones válidas
                                elegido = random.choice(poss)
                                brk[nm] = elegido.strftime('%H:%M')
                            else:
                                # Si no hay opciones en el rango, usamos la mediana
                                med = horas[len(horas)//2]
                                brk[nm] = med.strftime('%H:%M')
                    pivot[colb] = pivot['Nombre'].map(brk).fillna('')

                cols = ['DNI', 'Nombre', 'SERVICIO', 'SUPERIOR', 'Intervalo'] + \
                       list(DAY_NAMES.values()) + break_cols

                # Finalmente, volcamos el DataFrame completo en la hoja de Excel
                pivot[cols].to_excel(writer, sheet_name=hoja, index=False)
                ws = writer.sheets[hoja]

                # Coloreamos la fila de cabecera
                for idx, col in enumerate(cols):
                    cell = ws.cell(row=1, column=idx + 1)
                    if col in DAY_NAMES.values():
                        cell.fill = header_yellow
                    else:
                        cell.fill = header_blue
                    cell.font = font_bold

        # Guardamos el nombre del archivo en sesión y preparamos la URL de descarga
        session['last_file'] = file_name
        download_url = url_for('conversor.download')

    return render_template(
        'conversor.html',
        title='Conversor',
        services=SERVICES,
        download_url=download_url
    )


@conversor_bp.route('/conversor/download')
def download():
    upload_dir = current_app.config.get('UPLOAD_FOLDER', os.getcwd())
    last_file  = session.get('last_file')

    if last_file and os.path.exists(os.path.join(upload_dir, last_file)):
        path = os.path.join(upload_dir, last_file)
    else:
        # Si no hay last_file en sesión, buscamos el más reciente
        files = [
            f for f in os.listdir(upload_dir)
            if f.startswith("convertido_tabs_") and f.endswith(".xlsx")
        ]
        files = sorted(
            files,
            key=lambda f: os.path.getmtime(os.path.join(upload_dir, f)),
            reverse=True
        )
        if not files:
            flash("No se encontró ningún archivo para descargar.", "danger")
            return redirect(url_for('conversor.conversor'))
        path = os.path.join(upload_dir, files[0])

    return send_file(path, as_attachment=True, download_name=os.path.basename(path))
