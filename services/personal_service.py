import os
import math
import re
import pandas as pd
from datetime import datetime, timedelta, time, date
from calendar import monthrange
from openpyxl.styles import PatternFill


class PersonalService:
    """
    Servicio para procesar la nómina y generar el reporte de programación para Personal.
    Si se pasa 'servicio' igual a uno de los elementos en SERVICIOS_DISPONIBLES, genera
    un Excel con las 5 hojas (Nomina, Simulacion, Movimientos,
    Simulacion_Escalonada_Sugerida, Movimientos_Escalonados) solo para ese servicio.

    Si 'servicio' no coincide con ninguno de los elementos en SERVICIOS_DISPONIBLES,
    asume que debe procesar “todos los servicios”: recorre cada hoja de la lista
    SERVICIOS_DISPONIBLES y vuelca sus salidas en un único Excel consolidador.
    Retorna la ruta al archivo .xlsx resultante.
    """

    CONTRACT_HOURS = {'24HS': 6, '30HS': 6, '35HS': 7, '36HS': 6}

    # Mapeo nombre de hoja en “Requeridos” → clave interna para filtrar nómina
    SERVICE_KEY_MAP = {
        'Sop_Conectividad': 'Internet',
        'Sop_Flow':         'Flow',
        'Esp_CATV':         'CATV',
        'Esp_Movil':        'Movil',
        'Esp_XDSL':         'XDSL',
        'Digital':          'Digital',
        'CBS':              'CBS|PTF'
    }

    # Lista exacta de servicios individuales (coincide con las hojas disponibles en el Excel de Requeridos)
    SERVICIOS_DISPONIBLES = [
        "Sop_Conectividad", "Sop_Flow", "Esp_CATV",
        "Esp_Movil", "Esp_XDSL", "Digital", "CBS"
    ]

    INGRESOS_VALIDOS = [8, 9, 10, 11, 14, 15, 18, 19]

    def procesar(self, nomina_path: str, req_file, servicio: str, periodo: str) -> str:
        """
        Parámetros:
          - nomina_path: ruta al archivo .xlsx de nómina.
          - req_file: archivo binario .xlsx con hojas de requeridos.
          - servicio: nombre de la hoja a procesar (p.ej. "Sop_Conectividad").
                      Si este valor NO está en SERVICIOS_DISPONIBLES, se genera
                      un Excel único con la programación de todos los servicios listados en SERVICIOS_DISPONIBLES.
          - periodo: "mes", "sem1", "sem2", "sem3" o "sem4".
        Retorna la ruta absoluta al Excel resultante (.xlsx).
        """

        base = os.path.abspath(os.path.dirname(__file__))
        clave = servicio.strip()
        procesar_unico = (clave in self.SERVICIOS_DISPONIBLES)

        # Definimos el nombre de salida:
        if procesar_unico:
            out_filename = f"{clave}_reporte.xlsx"
        else:
            out_filename = "TodosServicios_reporte.xlsx"

        out_path = os.path.join(base, out_filename)

        # 1) Guardar temporalmente el archivo de requeridos
        temp_req_path = os.path.join(base, "temp_requeridos.xlsx")
        req_file.save(temp_req_path)

        def generar_por_servicio(svc: str, periodo: str):
            """
            Procesa la nómina + hoja 'svc' de 'temp_requeridos.xlsx' aplicando el filtro
            de 'periodo' y retorna un diccionario con los 5 DataFrames de salida.
            """
            # --- 1.1) Leer nómina ---
            df_nom = pd.read_excel(nomina_path)
            df_nom.columns = df_nom.columns.str.strip()

            # --- 1.2) Leer hoja 'svc' de requeridos y convertir a formato largo ---
            df_d = pd.read_excel(temp_req_path, sheet_name=svc, skiprows=[0, 2], header=0)
            df_d.rename(columns={df_d.columns[0]: 'Intervalo'}, inplace=True)

            # Convertir 'Intervalo' a tipo time
            df_d['Intervalo'] = pd.to_datetime(
                df_d['Intervalo'], format='%H:%M:%S', errors='coerce'
            ).dt.time
            df_d.dropna(subset=['Intervalo'], inplace=True)

            date_cols = [
                c for c in df_d.columns[1:]
                if not pd.isna(pd.to_datetime(str(c), errors='coerce'))
            ]
            df_long = (
                df_d.melt(
                    id_vars=['Intervalo'], value_vars=date_cols,
                    var_name='Fecha', value_name='Requeridos'
                )
                .dropna(subset=['Requeridos'])
            )
            df_long['Fecha'] = pd.to_datetime(
                df_long['Fecha'], format='%Y-%m-%d', errors='coerce'
            ).dt.date
            df_long['Requeridos'] = df_long['Requeridos'].astype(int)

            # --- Aplicar filtro por 'periodo' ---
            # Si no hay filas válidas, retornamos vacíos
            if df_long.empty:
                return {
                    "Nomina": pd.DataFrame(),
                    "Simulacion": pd.DataFrame(),
                    "Movimientos": pd.DataFrame(),
                    "Simulacion_Escalonada_Sugerida": pd.DataFrame(),
                    "Movimientos_Escalonados": pd.DataFrame()
                }

            # Determinar año y mes a partir de la primera fecha en df_long
            primer_fecha = df_long['Fecha'].min()
            anio = primer_fecha.year
            mes = primer_fecha.month
            dias_en_mes = monthrange(anio, mes)[1]  # e.g. 30 o 31

            # Rangos de cada semana
            sem1_inicio = date(anio, mes, 1)
            sem1_fin    = date(anio, mes, min(7, dias_en_mes))
            sem2_inicio = date(anio, mes, 8) if dias_en_mes >= 8 else sem1_fin
            sem2_fin    = date(anio, mes, min(14, dias_en_mes))
            sem3_inicio = date(anio, mes, 15) if dias_en_mes >= 15 else sem2_fin
            sem3_fin    = date(anio, mes, min(21, dias_en_mes))
            sem4_inicio = date(anio, mes, 22) if dias_en_mes >= 22 else sem3_fin
            sem4_fin    = date(anio, mes, dias_en_mes)

            if periodo == 'sem1':
                df_long = df_long[
                    (df_long['Fecha'] >= sem1_inicio) & (df_long['Fecha'] <= sem1_fin)
                ]
            elif periodo == 'sem2':
                df_long = df_long[
                    (df_long['Fecha'] >= sem2_inicio) & (df_long['Fecha'] <= sem2_fin)
                ]
            elif periodo == 'sem3':
                df_long = df_long[
                    (df_long['Fecha'] >= sem3_inicio) & (df_long['Fecha'] <= sem3_fin)
                ]
            elif periodo == 'sem4':
                df_long = df_long[
                    (df_long['Fecha'] >= sem4_inicio) & (df_long['Fecha'] <= sem4_fin)
                ]
            # Si periodo == 'mes', no filtramos

            # --- 1.3) Filtrar personal activo Y calcular INGRESO/EGRESO ---
            key = self.SERVICE_KEY_MAP.get(svc, svc)
            df_x = df_nom[
                df_nom['SERVICIO'].str.contains(key, case=False, na=False) &
                (df_nom['ACTIVO'].str.upper() == 'ACTIVO')
            ].copy()

            # Convertir 'INGRESO' a tipo time
            df_x['INGRESO'] = pd.to_datetime(
                df_x['INGRESO'].astype(str),
                format='%H:%M:%S',
                errors='coerce'
            ).dt.time

            # Calcular EGRESO sumándole horas según contrato
            df_x['EGRESO'] = [
                (datetime.combine(datetime.today(), ing) +
                 timedelta(hours=self.CONTRACT_HOURS.get(str(con).strip().upper(), 24))
                ).time()
                for ing, con in zip(df_x['INGRESO'], df_x['CONTRATO'])
            ]
            df_x = df_x.sort_values('NOMBRE').reset_index(drop=True)

            # --- 1.4) Asignar off days dinámicos ---
            def assign_off_days(df_tmp):
                offs = []
                weekend_idxs = [
                    i for i, r in df_tmp.iterrows()
                    if str(r['CONTRATO']).strip().upper() in ('30HS', '35HS')
                ]
                half = len(weekend_idxs) // 2
                for i, row in df_tmp.iterrows():
                    c = str(row['CONTRATO']).strip().upper()
                    if c == '24HS':
                        offs.append([(i + k) % 7 for k in range(3)])
                    elif c in ('30HS', '35HS'):
                        wd = i % 5
                        we = 5 if weekend_idxs.index(i) < half else 6
                        offs.append([wd, we])
                    elif c == '36HS':
                        offs.append([5 if i % 2 == 0 else 6])
                    else:
                        offs.append([])
                return offs

            df_x['OFF_DAYS'] = assign_off_days(df_x)
            def is_off(row, f): return f.weekday() in row['OFF_DAYS']

            rows = []
            simulacion = []
            movimientos = []

            # --- 1.5) Bucle principal: asignación para cada (Fecha, Intervalo) ---
            for _, r in df_long.iterrows():
                f = r['Fecha']
                i = r['Intervalo']
                req = r['Requeridos']

                # Límites dinámicos
                if req < 10:
                    li, up = max(req - 1, 0), req + 1
                elif req < 20:
                    li, up = max(req - 2, 0), req + 2
                else:
                    li, up = math.floor(req * 0.9), math.ceil(req * 1.1)

                prime = 'Prime' if time(9, 0) <= i < time(21, 0) else 'No prime'

                # Filtrar quienes NO están de franco
                df_av = df_x[~df_x.apply(lambda row: is_off(row, f), axis=1)]
                norm = (df_av['INGRESO'] <= i) & (i < df_av['EGRESO'])
                wrap = (df_av['EGRESO'] < df_av['INGRESO']) & (
                    (df_av['INGRESO'] <= i) | (i < df_av['EGRESO'])
                )
                pres = df_av[norm | wrap].copy()
                cnt = len(pres)

                # Regla especial para domingo (36HS)
                if f.weekday() == 6:
                    sab = f - timedelta(days=1)
                    hora = i.strftime('%H:%M')
                    used = {
                        nm
                        for rec in rows
                        for nm in rec['Nombres_Presentes'].split(';')
                        if rec['Fecha'] == sab and rec['Intervalo'] == hora
                    }
                    pres = pres[~pres['NOMBRE'].isin(used)]
                    p36 = pres[pres['CONTRATO'].str.upper() == '36HS']
                    oth = pres.drop(p36.index)
                    need = max(li - len(p36), 0)
                    pres = pd.concat([p36, oth.head(need)])
                    cnt = len(pres)

                falt = max(li - cnt, 0)
                sobr = max(cnt - up, 0)
                if falt > 0:
                    movimientos.append({
                        'Fecha': f,
                        'Intervalo': i.strftime('%H:%M'),
                        'Mover': falt,
                        'Desde': '',
                        'Hacia': ''
                    })

                leader_col = next(
                    (c for c in pres.columns if c.strip().lower() in ('superior', 'jefe', 'lider')),
                    None
                )
                lideres = pres[leader_col].dropna().unique().tolist() if leader_col else []
                estado = (
                    'UNDER' if cnt < li else
                    'OVER'  if cnt > up else
                    'LIMITE' if cnt == li else
                    'OK'
                )

                rec = {
                    'Fecha': f,
                    'Intervalo': i.strftime('%H:%M'),
                    'Prime': prime,
                    'Requeridos': req,
                    'Limite Inferior': li,
                    'Limite Superior': up,
                    'Faltante': falt,
                    'Sobrantes': sobr,
                    'Asignados': cnt,
                    'Estado': estado,
                    'Lider': ';'.join(lideres),
                    'Movimientos': '',
                    'Nombres_Presentes': ';'.join(pres['NOMBRE'].astype(str).unique())
                }
                rows.append(rec)
                simulacion.append(rec.copy())

            # --- 1.6) Calcular “Desde” y “Hacia” para cada movimiento ---
            updated = []
            for mov in movimientos:
                date0 = mov['Fecha']
                int_tm = datetime.strptime(mov['Intervalo'], '%H:%M').time()
                search_date = date0 - timedelta(days=1) if int_tm < time(1, 0) else date0
                base_dt = datetime.combine(search_date, int_tm)

                donors = []
                for r in rows:
                    if r['Estado'] == 'OVER' and r['Fecha'] == search_date:
                        cand_dt = datetime.combine(
                            search_date,
                            datetime.strptime(r['Intervalo'], '%H:%M').time()
                        )
                        delta = (cand_dt - base_dt).total_seconds()
                        if -2*3600 <= delta <= 2*3600 and delta != 0:
                            donors.append((abs(delta), r['Intervalo']))
                if donors:
                    raw_desde = min(donors, key=lambda x: x[0])[1]
                else:
                    eve = []
                    for r in rows:
                        if r['Estado'] == 'OVER' and r['Fecha'] == search_date:
                            ct = datetime.strptime(r['Intervalo'], '%H:%M').time()
                            cd = datetime.combine(search_date, ct)
                            if cd <= datetime.combine(search_date, time(18, 30)):
                                dist = abs((datetime.combine(search_date, time(19, 0)) - cd).total_seconds())
                                eve.append((dist, r['Intervalo']))
                    raw_desde = eve and min(eve, key=lambda x: x[0])[1] or '19:00 (extraordinario)'

                if '(' in raw_desde:
                    mov['Desde'] = raw_desde
                else:
                    hh = int(raw_desde.split(':')[0])
                    cands = [h for h in self.INGRESOS_VALIDOS if h <= hh]
                    sel = max(cands) if cands else min(self.INGRESOS_VALIDOS)
                    mov['Desde'] = f"{sel:02d}:00"

                if int_tm < time(1, 0):
                    mov['Hacia'] = '19:00'
                else:
                    hh = int(int_tm.hour + (1 if int_tm.minute > 0 else 0))
                    cands = [h for h in self.INGRESOS_VALIDOS if h >= hh]
                    sel = min(cands) if cands else max(self.INGRESOS_VALIDOS)
                    mov['Hacia'] = f"{sel:02d}:00"

                updated.append(mov)
            movimientos = updated

            # --- 1.7) Aplicar movimientos en simulación “normal” ---
            for mov in movimientos:
                for rec in simulacion:
                    if rec['Fecha'] == mov['Fecha'] and rec['Intervalo'] == mov['Intervalo']:
                        rec['Asignados'] += mov['Mover']
                        rec['Estado'] = 'LIMITE'
                        rec['Movimientos'] = f"{mov['Mover']} desde {mov['Desde']} → {mov['Hacia']}"

            # --- 1.8) Generar hoja “Simulacion_Escalonada_Sugerida” ---
            df_sim = pd.DataFrame(simulacion)
            df_sim['Fecha'] = pd.to_datetime(df_sim['Fecha'], format='%Y-%m-%d')

            df_sim_sugerida = df_sim.copy()
            df_sim_sugerida['Escalona_Sugerida'] = ""

            pattern = re.compile(r'(\d+)\s+desde\s+(\d{2}:\d{2})\s+→\s+(\d{2}:\d{2})')

            def generar_sugerencia(movimientos_str: str) -> str:
                if not movimientos_str or not isinstance(movimientos_str, str):
                    return ""
                sugerencias = []
                partes = [p.strip() for p in movimientos_str.split(';') if p.strip()]
                for parte in partes:
                    m = pattern.match(parte)
                    if not m:
                        continue
                    cantidad = int(m.group(1))
                    hora_inicio = m.group(2)
                    hora_fin = m.group(3)

                    t_inicio = datetime.strptime(hora_inicio, "%H:%M")
                    t_fin = datetime.strptime(hora_fin, "%H:%M")
                    diff_horas = int((t_fin - t_inicio).total_seconds() // 3600)
                    if diff_horas < 0:
                        continue
                    if diff_horas > 2:
                        current = t_inicio
                        for _ in range(diff_horas):
                            siguiente = current + timedelta(hours=1)
                            sugerencias.append(
                                f"{cantidad} desde {current.strftime('%H:%M')} → {siguiente.strftime('%H:%M')}"
                            )
                            current = siguiente
                return "; ".join(sugerencias)

            df_sim_sugerida['Escalona_Sugerida'] = df_sim_sugerida['Movimientos'].apply(generar_sugerencia)

            # --- 1.9) Preparar DataFrames de salida para este servicio ---
            df_nomina_out = pd.DataFrame(rows).drop(columns=['Movimientos'])
            df_sim_out = df_sim.copy()
            df_mov_out = pd.DataFrame(movimientos)[['Fecha', 'Intervalo', 'Mover', 'Desde', 'Hacia']]

            # Movimientos escalonados sugeridos
            movimientos_escalonados = []
            for _, rec in df_sim_sugerida.iterrows():
                if rec['Escalona_Sugerida']:
                    partes = [p.strip() for p in rec['Escalona_Sugerida'].split(';') if p.strip()]
                    for parte in partes:
                        m = pattern.match(parte)
                        if not m:
                            continue
                        cantidad = int(m.group(1))
                        inicio = m.group(2)
                        fin = m.group(3)
                        movimientos_escalonados.append({
                            'Fecha': rec['Fecha'].date(),
                            'Mover': cantidad,
                            'Desde': inicio,
                            'Hacia': fin
                        })
            df_mov_esc_out = pd.DataFrame(movimientos_escalonados)

            return {
                "Nomina": df_nomina_out,
                "Simulacion": df_sim_out,
                "Movimientos": df_mov_out,
                "Simulacion_Escalonada_Sugerida": df_sim_sugerida,
                "Movimientos_Escalonados": df_mov_esc_out
            }

        # --------------------------------------------------
        # 2) Decidir si procesar UN servicio o TODOS
        # --------------------------------------------------
        if procesar_unico:
            # Caso 1: 'servicio' coincide con un nombre real de SERVICIOS_DISPONIBLES
            outputs = generar_por_servicio(clave, periodo)
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                # -----------------------------------------------------------------------------------
                # Hoja “Nomina”
                df_nomina = outputs["Nomina"]
                hoja_nom = "Nomina"
                df_nomina.to_excel(writer, sheet_name=hoja_nom, index=False)
                ws = writer.sheets[hoja_nom]
                if not df_nomina.empty:
                    cols = df_nomina.columns.tolist()
                    li_idx = cols.index("Limite Inferior") + 1
                    up_idx = cols.index("Limite Superior") + 1
                    asig_idx = cols.index("Asignados") + 1
                    est_idx = cols.index("Estado") + 1
                    for i in range(2, ws.max_row + 1):
                        v = ws.cell(i, asig_idx).value
                        l = ws.cell(i, li_idx).value
                        u = ws.cell(i, up_idx).value
                        c = ws.cell(i, est_idx)
                        c.fill = (
                            PatternFill('solid', fgColor='FF0000') if v < l else
                            PatternFill('solid', fgColor='FFFF00') if v > u else
                            PatternFill('solid', fgColor='FFA500') if v == l else
                            PatternFill('solid', fgColor='00FF00')
                        )

                # -----------------------------------------------------------------------------------
                # Hoja “Simulacion”
                df_sim = outputs["Simulacion"]
                hoja_sim = "Simulacion"
                df_sim.to_excel(writer, sheet_name=hoja_sim, index=False)
                ws = writer.sheets[hoja_sim]
                if not df_sim.empty:
                    cols = df_sim.columns.tolist()
                    li_idx = cols.index("Limite Inferior") + 1
                    up_idx = cols.index("Limite Superior") + 1
                    asig_idx = cols.index("Asignados") + 1
                    est_idx = cols.index("Estado") + 1
                    for i in range(2, ws.max_row + 1):
                        v = ws.cell(i, asig_idx).value
                        l = ws.cell(i, li_idx).value
                        u = ws.cell(i, up_idx).value
                        c = ws.cell(i, est_idx)
                        c.fill = (
                            PatternFill('solid', fgColor='FF0000') if v < l else
                            PatternFill('solid', fgColor='FFFF00') if v > u else
                            PatternFill('solid', fgColor='FFA500') if v == l else
                            PatternFill('solid', fgColor='00FF00')
                        )

                # -----------------------------------------------------------------------------------
                # Hoja “Movimientos”
                df_mov = outputs["Movimientos"]
                if not df_mov.empty:
                    hoja_mov = "Movimientos"
                    df_mov.to_excel(writer, sheet_name=hoja_mov, index=False)

                # -----------------------------------------------------------------------------------
                # Hoja “Simulacion_Escalonada_Sugerida”
                df_sug = outputs["Simulacion_Escalonada_Sugerida"]
                hoja_sug = "Sim_EscSug"
                df_sug.to_excel(writer, sheet_name=hoja_sug, index=False)
                ws = writer.sheets[hoja_sug]
                if not df_sug.empty:
                    cols = df_sug.columns.tolist()
                    li_idx = cols.index("Limite Inferior") + 1
                    up_idx = cols.index("Limite Superior") + 1
                    asig_idx = cols.index("Asignados") + 1
                    est_idx = cols.index("Estado") + 1
                    for i in range(2, ws.max_row + 1):
                        v = ws.cell(i, asig_idx).value
                        l = ws.cell(i, li_idx).value
                        u = ws.cell(i, up_idx).value
                        c = ws.cell(i, est_idx)
                        c.fill = (
                            PatternFill('solid', fgColor='FF0000') if v < l else
                            PatternFill('solid', fgColor='FFFF00') if v > u else
                            PatternFill('solid', fgColor='FFA500') if v == l else
                            PatternFill('solid', fgColor='00FF00')
                        )

                # -----------------------------------------------------------------------------------
                # Hoja “Movimientos_Escalonados”
                df_mov_esc = outputs["Movimientos_Escalonados"]
                if not df_mov_esc.empty:
                    hoja_esc_mov = "Mov_Escalonados"
                    df_mov_esc.to_excel(writer, sheet_name=hoja_esc_mov, index=False)

            return out_path

        else:
            # Caso “TODOS los servicios”: recorro cada servicio real y vuelco sus 5 hojas
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                for svc_nombre in self.SERVICIOS_DISPONIBLES:
                    outputs = generar_por_servicio(svc_nombre, periodo)

                    # ------------------------------------------------------------------------------------------------
                    # Hoja “<svc_nombre>_Nomina”
                    df_nomina = outputs["Nomina"]
                    hoja_nom = f"{svc_nombre[:12]}_Nomina"
                    df_nomina.to_excel(writer, sheet_name=hoja_nom, index=False)
                    ws = writer.sheets[hoja_nom]
                    if not df_nomina.empty:
                        cols = df_nomina.columns.tolist()
                        li_idx = cols.index("Limite Inferior") + 1
                        up_idx = cols.index("Limite Superior") + 1
                        asig_idx = cols.index("Asignados") + 1
                        est_idx = cols.index("Estado") + 1
                        for i in range(2, ws.max_row + 1):
                            v = ws.cell(i, asig_idx).value
                            l = ws.cell(i, li_idx).value
                            u = ws.cell(i, up_idx).value
                            c = ws.cell(i, est_idx)
                            c.fill = (
                                PatternFill('solid', fgColor='FF0000') if v < l else
                                PatternFill('solid', fgColor='FFFF00') if v > u else
                                PatternFill('solid', fgColor='FFA500') if v == l else
                                PatternFill('solid', fgColor='00FF00')
                            )

                    # ------------------------------------------------------------------------------------------------
                    # Hoja “<svc_nombre>_Simulacion”
                    df_sim = outputs["Simulacion"]
                    hoja_sim = f"{svc_nombre[:12]}_Simulacion"
                    df_sim.to_excel(writer, sheet_name=hoja_sim, index=False)
                    ws = writer.sheets[hoja_sim]
                    if not df_sim.empty:
                        cols = df_sim.columns.tolist()
                        li_idx = cols.index("Limite Inferior") + 1
                        up_idx = cols.index("Limite Superior") + 1
                        asig_idx = cols.index("Asignados") + 1
                        est_idx = cols.index("Estado") + 1
                        for i in range(2, ws.max_row + 1):
                            v = ws.cell(i, asig_idx).value
                            l = ws.cell(i, li_idx).value
                            u = ws.cell(i, up_idx).value
                            c = ws.cell(i, est_idx)
                            c.fill = (
                                PatternFill('solid', fgColor='FF0000') if v < l else
                                PatternFill('solid', fgColor='FFFF00') if v > u else
                                PatternFill('solid', fgColor='FFA500') if v == l else
                                PatternFill('solid', fgColor='00FF00')
                            )

                    # ------------------------------------------------------------------------------------------------
                    # Hoja “<svc_nombre>_Movimientos”
                    df_mov = outputs["Movimientos"]
                    if not df_mov.empty:
                        hoja_mov = f"{svc_nombre[:12]}_Movimientos"
                        df_mov.to_excel(writer, sheet_name=hoja_mov, index=False)

                    # ------------------------------------------------------------------------------------------------
                    # Hoja “<svc_nombre>_Simulacion_Escalonada_Sugerida”
                    df_sug = outputs["Simulacion_Escalonada_Sugerida"]
                    hoja_sug = f"{svc_nombre[:12]}_SimEscSug"
                    df_sug.to_excel(writer, sheet_name=hoja_sug, index=False)
                    ws = writer.sheets[hoja_sug]
                    if not df_sug.empty:
                        cols = df_sug.columns.tolist()
                        li_idx = cols.index("Limite Inferior") + 1
                        up_idx = cols.index("Limite Superior") + 1
                        asig_idx = cols.index("Asignados") + 1
                        est_idx = cols.index("Estado") + 1
                        for i in range(2, ws.max_row + 1):
                            v = ws.cell(i, asig_idx).value
                            l = ws.cell(i, li_idx).value
                            u = ws.cell(i, up_idx).value
                            c = ws.cell(i, est_idx)
                            c.fill = (
                                PatternFill('solid', fgColor='FF0000') if v < l else
                                PatternFill('solid', fgColor='FFFF00') if v > u else
                                PatternFill('solid', fgColor='FFA500') if v == l else
                                PatternFill('solid', fgColor='00FF00')
                            )

                    # ------------------------------------------------------------------------------------------------
                    # Hoja “<svc_nombre>_Movimientos_Escalonados”
                    df_mov_esc = outputs["Movimientos_Escalonados"]
                    if not df_mov_esc.empty:
                        hoja_esc_mov = f"{svc_nombre[:12]}_Mov_Escalonados"
                        df_mov_esc.to_excel(writer, sheet_name=hoja_esc_mov, index=False)

            return out_path
